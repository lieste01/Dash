import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook

def matrixCalor(data, jana):
    #dataFrame = pd.read_excel(f"{os.getcwd()}\\Temp Enr.xlsx")
    dataFrame = pd.read_excel(f"Temp Enr.xlsx")

    nameColumns = ['Data']

    # Cria uma array com os novos nomes das colunas
    for usina in range(1, 21):
        for ts in range(1, 8):
            nameColumns.append(f"Jana {usina} - TS {ts}")

    # Muda o nome das colunas, substitution com os nomes da array nameColumns
    dataFrame.columns = nameColumns

    # Transforma a coluna Data em Date
    dataFrame['Data'] = pd.to_datetime(dataFrame['Data']).dt.strftime('%Y-%m-%d')

    # Substitui os valores [-11059] No Good Data For Calculation para 0
    dataFrame = dataFrame.replace('[-11059] No Good Data For Calculation', 0)
    dataFrame = dataFrame.replace('[-11057] Not Enough Values For Calculation', 0)
    # Filtro usando indexação booleana
    df_filtered = dataFrame[(dataFrame['Data'] == f"{data}")][
        ['Data', f'{jana} - TS 1', f'{jana} - TS 2', f'{jana} - TS 3',
         f'{jana} - TS 4', f'{jana} - TS 5', f'{jana} - TS 6',
         f'{jana} - TS 7']]
    colunas = [f'{jana} - TS 1', f'{jana} - TS 2', f'{jana} - TS 3',
               f'{jana} - TS 4', f'{jana} - TS 5', f'{jana} - TS 6',
               f'{jana} - TS 7']

    # Defina o tamanho da matriz (número de linhas e colunas)
    matrix = np.zeros((len(df_filtered.columns), len(df_filtered.columns)))

    for row, col in df_filtered.iterrows():
        for num in range(1, len(df_filtered.columns)):
            for num2 in range(1, len(df_filtered.columns)):
                if col[num] == 0:
                    matrix[num][num2] = 0
                elif col[num2] == 0:
                    matrix[num][num2] = 0
                else:
                    matrix[num][num2] = round((float(col[num]) / float(col[num2])) - 1, 2)

    matrix = [linha[1:] for linha in matrix[1:]]

    # Converter a matriz para um dataframe para que as datas sejam os rótulos das linhas e colunas
    matrix_df = pd.DataFrame(matrix, columns=colunas, index=colunas)

    return matrix_df

janas = ['Jana 15', 'Jana 16', 'Jana 17', 'Jana 18', 'Jana 19', 'Jana 20']

dataFrame = pd.read_excel(f"Temp Enr.xlsx")
matrixCrunch = []
for data in dataFrame["Data"].dt.strftime("%Y-%m-%d"):
    for jana in janas:
        matrix = matrixCalor(data, jana)
        matrix["Data"] = data
        matrixCrunch.append(matrix)


dtCrunch = [pd.DataFrame(layer) for layer in matrixCrunch]



# Criar um novo arquivo Excel
wb = Workbook()

# Adicionar cada DataFrame como uma aba no arquivo Excel
for idx, df in enumerate(dtCrunch):
    lista = df.columns[0]
    sheet = wb.create_sheet(title=f"{df['Data'][0][5:10]}-{lista[0][:7]}")

    # Preencher os nomes das colunas a partir da coluna 2
    for c_idx, col_name in enumerate(df.columns, start=2):
        sheet.cell(row=1, column=c_idx, value=col_name)

    # Preencher os nomes das linhas e valores
    for r_idx, (row_name, row_data) in enumerate(df.iterrows(), start=2):
        sheet.cell(row=r_idx, column=1, value=row_name)
        for c_idx, value in enumerate(row_data, start=2):
            sheet.cell(row=r_idx, column=c_idx, value=value)

# Remover a aba de planilha padrão
default_sheet = wb.get_sheet_by_name('Sheet')
wb.remove_sheet(default_sheet)

# Salvar o arquivo Excel
wb.save('dataframes.xlsx')

print("Exportação concluída.")